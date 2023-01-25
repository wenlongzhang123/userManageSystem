from django.shortcuts import render, redirect
from app01 import models
from openpyxl import load_workbook

def depart_list(request):
    # 部门列表
    # if request.method == "GET":
    queryset = models.Department.objects.all()
    return render(request, "depart_list.html", {"queryset": queryset})

    # return render(request, "depart_add.html")

def depart_add(request):
    if request.method == "GET":
        return render(request, "depart_add.html")
    title = request.POST.get("a")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

def depart_edit(request, nid):
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"row_object": row_object})
    title = request.POST.get("b")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")

def depart_multi(request):
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl,由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):  # 从文件的第二行开始读取
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/depart/list/")